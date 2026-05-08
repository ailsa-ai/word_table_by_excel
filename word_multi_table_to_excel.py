from typing import Optional
import os  # 操作系统库：用来遍历文件夹、创建文件夹、拼接文件路径、判断文件是否存在。
import re
from docx import Document
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter  # 把列数字 1,2,3 转成 Excel 的 A,B,C，用来自动适配列宽。


# 清洗 Sheet 工作表名称
def clean_sheet_name(name: str) -> str:
    invalid_chars = ['\\', '/', ':', '*', '?', '[', ']']
    for char in invalid_chars:
        name = name.replace(char, "")
    return name[:31].strip() or "Sheet"


# 占位符标黄
def highlight_placeholders(ws):
    placeholder_pattern = re.compile(r"#.+?#")
    fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and placeholder_pattern.search(str(cell.value)):
                cell.fill = fill


# 自动适配 Excel 列宽
def auto_adjust_column_width(ws):
    for col in ws.columns:
        max_length = 0
        column_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                cell_text = str(cell.value) if cell.value else ""
                line_lengths = [len(line) for line in cell_text.splitlines()]
                if line_lengths:
                    max_length = max(max_length, max(line_lengths))
            except:
                pass
        adjusted_width = min(max_length + 2, 100)
        ws.column_dimensions[column_letter].width = adjusted_width


# Excel 统一样式设置（边框、居中、加粗、底色）
def set_cell_style(ws, total_cols: int):
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

    # 合并第一行标题并居中
    if total_cols > 1:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
        first_cell = ws.cell(row=1, column=1)
        first_cell.alignment = Alignment(horizontal='center', vertical='center')
        first_cell.font = Font(bold=True)
        first_cell.fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")

    # 其他单元格设置样式
    for row_idx, row in enumerate(ws.iter_rows(min_row=2)):
        for cell in row:
            cell.alignment = alignment
            cell.border = thin_border
            # 第二行（表头）加粗居中
            if row_idx == 0:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")


# 【最核心】智能识别标题函数
def get_table_smart_title(doc, table_index: int, table) -> str:
    """
    智能判断标题来源：
    1. 优先取表格上方的段落标题（过滤序号）
    2. 如果没有，再取表格第一行的去重内容
    """
    # 1. 先找表格上方的段落标题
    content_list = []
    for elem in doc.element.body:
        if elem.tag.endswith('p'):
            content_list.append(('para', elem))
        elif elem.tag.endswith('tbl'):
            content_list.append(('table', elem))

    table_element = table._tbl
    table_pos = None
    for idx, (type_, elem) in enumerate(content_list):
        if type_ == 'table' and elem == table_element:
            table_pos = idx
            break

    if table_pos is not None:
        for i in range(table_pos - 1, -1, -1):
            type_, elem = content_list[i]
            if type_ == 'para':
                para_text = elem.text.strip() if elem.text else ""
                # 过滤纯序号行（如"1."、"2."）
                if para_text and len(para_text) < 60 and not para_text.endswith('.'):
                    return para_text

    # 2. 上方没有标题，取表格第一行去重内容
    if len(table.rows) > 0:
        first_row_cells = [cell.text.strip() for cell in table.rows[0].cells if cell.text.strip()]
        unique_cells = []
        seen = set()
        for cell_text in first_row_cells:
            if cell_text not in seen:
                seen.add(cell_text)
                unique_cells.append(cell_text)
        row_title = " ".join(unique_cells).strip()
        if row_title:
            return row_title

    # 兜底
    return f"表格{table_index + 1}"


# 单个 Word 转 Excel 主函数
def word_tables_to_excel_by_sheet(word_path, excel_path):
    try:
        if not os.path.exists(word_path):
            print(f"文件不存在：{os.path.basename(word_path)}")
            return False

        doc = Document(word_path)
        tables = doc.tables
        if not tables:
            print(f"{os.path.basename(word_path)} 未找到任何表格")
            return False

        wb = Workbook()
        wb.remove(wb.active)

        for table_idx, table in enumerate(tables):
            if len(table.rows) == 0:
                continue

            # 智能获取标题（优先上方段落，再取表格第一行）
            table_title = get_table_smart_title(doc, table_idx, table)
            sheet_name = clean_sheet_name(table_title)
            ws = wb.create_sheet(title=sheet_name)

            total_cols = len(table.columns)
            # 写入标题到Excel第一行（合并居中）
            ws.cell(row=1, column=1, value=table_title)

            # 写入表格数据：
            # 情况1：如果标题来自表格上方，写入所有行
            # 情况2：如果标题来自表格第一行，跳过第一行避免重复
            if len(table.rows) > 0:
                # 判断标题是否来自表格第一行
                first_row_cells = [cell.text.strip() for cell in table.rows[0].cells if cell.text.strip()]
                first_row_text = " ".join(list(dict.fromkeys(first_row_cells))).strip()
                if table_title == first_row_text:
                    # 标题来自表格第一行，跳过第一行
                    for row_idx, row in enumerate(table.rows[1:], start=2):
                        for col_idx, cell in enumerate(row.cells, start=1):
                            cell_text = cell.text
                            ws.cell(row=row_idx, column=col_idx, value=cell_text)
                else:
                    # 标题来自上方段落，写入所有行
                    for row_idx, row in enumerate(table.rows, start=2):
                        for col_idx, cell in enumerate(row.cells, start=1):
                            cell_text = cell.text
                            ws.cell(row=row_idx, column=col_idx, value=cell_text)

            # 应用样式
            set_cell_style(ws, total_cols)
            auto_adjust_column_width(ws)
            highlight_placeholders(ws)

        wb.save(excel_path)
        print(f"转换成功：{os.path.basename(word_path)} → {len(tables)} 个Sheet")
        return True

    except PermissionError:
        print(f"权限错误：请关闭已打开的 Excel 文件：{os.path.basename(excel_path)}")
        return False
    except Exception as e:
        print(f"转换失败 {os.path.basename(word_path)}：{str(e)}")
        return False


# 批量处理整个文件夹
def batch_word_to_excel(folder_path):
    out_dir = os.path.join(folder_path, "Excel分表输出（双场景自动适配）")
    os.makedirs(out_dir, exist_ok=True)

    total_files = 0
    success_files = 0

    for filename in os.listdir(folder_path):
        if filename.startswith("~$") or not filename.lower().endswith(".docx"):
            continue

        total_files += 1
        word_full = os.path.join(folder_path, filename)
        excel_name = os.path.splitext(filename)[0] + ".xlsx"
        excel_full = os.path.join(out_dir, excel_name)

        if word_tables_to_excel_by_sheet(word_full, excel_full):
            success_files += 1

    print("\n" + "=" * 50)
    print(f"处理完成！共扫描 {total_files} 个Word，成功 {success_files} 个")
    print(f"输出目录：{out_dir}")
    print("=" * 50)


if __name__ == "__main__":
    current_folder = os.getcwd()
    print(f"正在处理当前目录所有Word文件：{current_folder}\n")
    batch_word_to_excel(current_folder)
    input("\n按回车键退出...")
